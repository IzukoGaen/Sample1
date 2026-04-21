"""Write QC workbook (openpyxl) from a ComparisonResult."""

from __future__ import annotations

import io

import openpyxl as py
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows

from sanitycheck.models import ComparisonResult


def _build_qc_workbook(result: ComparisonResult) -> py.Workbook:
    if result.tbl_overall_change is None or result.tbl_filenames is None:
        raise ValueError("ComparisonResult missing required tables for export.")

    wb = py.Workbook()
    dsp = NamedStyle(name="dsp")
    wb.add_named_style(dsp)
    ws = wb.active
    ft = Font(bold=True)
    ft_header = Font(bold=True, color="ffffff")
    next_header_row = 0

    def append_table(
        table,
        special: bool = False,
        header_color: str = "357ec7",
        end_column: str = "B",
    ) -> None:
        nonlocal next_header_row
        for r in dataframe_to_rows(table, index=False, header=True):
            ws.append(r)
        coord = "A{}:{}{}".format(next_header_row + 1, end_column, next_header_row + 1)
        for cell in ws[coord][0]:
            cell.font = ft_header
        for cell in ws[coord][0]:
            cell.fill = PatternFill("solid", fgColor=header_color)
        if special:
            for cell in ws[next_header_row + 2]:
                cell.font = ft
        ws.append([""])
        next_header_row += len(table) + 2

    def append_comment(comment: str, end_column: str = "B") -> None:
        nonlocal next_header_row
        ws["A{}".format(next_header_row + 1)].value = comment
        ws["A{}".format(next_header_row + 1)].font = Font(italic=True, color="990012")
        ws.append([""])
        next_header_row += 2

    def append_title(title: str, title_size: int = 14) -> None:
        nonlocal next_header_row
        ws.append([title])
        ws.append([""])
        for cell in ws[next_header_row + 1]:
            cell.font = Font(size=title_size, italic=True)
        next_header_row += 2

    indices_intersection = result.tbl_overall_change.set_index("Status").index.intersection(
        ["Modified"]
    )
    sheets_ = list(
        result.tbl_overall_change.set_index("Status").loc[indices_intersection, "Worksheet"].values
    )

    append_title("Sanity Check QC", 20)
    append_table(result.tbl_filenames, header_color="9f000f")

    append_title("Overall Change Summary")
    append_table(result.tbl_overall_change, header_color="348017")

    if "Data Feed Setup" in sheets_ and result.feed_differences is not None:
        append_title("Data Feed Setup")
        append_table(result.feed_differences, end_column="D")

    if "Subsidiary Mapping" in sheets_ and result.tbl_smry_val_comparison is not None:
        append_title("Subsidiary Mapping")
        append_table(result.tbl_smry_val_comparison, end_column="C")

    if "Screen Criteria" in sheets_ and result.tbl_screening_changes is not None:
        append_title("Screening Criteria")
        append_table(result.tbl_screening_changes, end_column="E")
        append_comment(
            comment=(
                "Row column indicates the Column Header's row number in its file of origin. "
                "For added Column Headers, it displays row number in Test file. For removed ones, "
                "row number comes from Feed file. All changing Column Headers have Test's row number."
            )
        )

    if "Factor List" in sheets_ and result.tbl_factor_list_merge is not None:
        append_title("Factor List")
        append_table(result.tbl_factor_list_merge, end_column="C")

    if "APL Definition" in sheets_:
        append_title("APL")
        if result.tbl_first_apl is not None and not result.tbl_first_apl.empty:
            append_table(result.tbl_first_apl, end_column="D")
        if result.tbl_second_apl_smry is not None and not result.tbl_second_apl_smry.empty:
            append_table(result.tbl_second_apl_smry, end_column="D")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40

    custom_style = DifferentialStyle(
        font=Font(bold=True, italic=True, color=Color("f70d1a"))
    )
    ruleone = Rule(type="expression", dxf=custom_style)
    ruleone.formula = ['$B2="Modified"']
    ruletwo = Rule(type="expression", dxf=custom_style)
    ruletwo.formula = ['$B2="Added"']
    rulethree = Rule(type="expression", dxf=custom_style)
    rulethree.formula = ['$B2="Removed"']
    ws.conditional_formatting.add("B2:B999", ruleone)
    ws.conditional_formatting.add("B2:B999", ruletwo)
    ws.conditional_formatting.add("B2:B999", rulethree)

    return wb


def export_log(result: ComparisonResult, log_name: str) -> None:
    wb = _build_qc_workbook(result)
    wb.save(log_name)


def export_log_bytes(result: ComparisonResult) -> bytes:
    """Build the same QC workbook and return it as ``.xlsx`` bytes (no temp file)."""
    wb = _build_qc_workbook(result)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
